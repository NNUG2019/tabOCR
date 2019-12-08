from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from random import choice


def generate_font(name, size, bold=False, italic=False):
    """ Generate Font object.
        :Arguments:
            name: str:: font name
            size: int:: font size
            bold: bool:: is font bold
            italic: bool: is font italic
        :Returns:
            Font object
    """
    return Font(name, size, bold, italic)


def generate_alingment(horizontal, vertical):
    """ Generate Alingment object.
        :Arguments:
            horizontal: type of the horizontal alingment
            vertical: type of the vertical alingment
        :Returns:
            Alingment object
    """
    return Alignment(horizontal, vertical)


def generate_pattern_fill(color, patternType='solid', fill_type='solid'):
    """ Generate PatternFill object.
        :Arguments:
            color: str:: hex code of the color
            patternType: str:: pattern type
            fill_type: str:: fill type
        :Returns:
            PatternFill object
    """
    return PatternFill(patternType=patternType, fill_type=fill_type,
                       fgColor=color)


def define_border(style, color):
    """ Generate Side object.
        :Arguments:
            style: str:: border style
            color: str:: hex code of the color
        :Returns:
            Side object
    """
    return Side(border_style=style, color=color)


def define_border_style(line_style, color="FFFFFF"):
    """ Define border for each cell side.
        :Arguments:
            line_style: str:: border line style (eg. dashed, thin, etc.)
            color: str:: hex code of the color
        :Returns:
            dict: defined border
    """
    return {"top": define_border(line_style, color),
            "bottom": define_border(line_style, color),
            "left": define_border(line_style, color),
            "right": define_border(line_style, color)}


def generate_border(line_style, is_border=False, is_partial_border=False):
    """ Generate Border object.
        :Arguments:
            line_style: str:: border line style (eg. dashed, thin, etc.)
            is_border: bool: is table has border
            is_partial_border: bool:: is table has partial border
                                      (only columns/only rows)
        :Returns:
            Border object
    """
    if is_border:
        style = define_border_style(line_style, color="000000")
        if is_partial_border:
            side = choice(["top", "left"])
            style = define_border_style(line_style)
            style = {**style, side: define_border(line_style, color="000000")}
        return Border(**style)
    else:
        return Border(**define_border_style(line_style, color="FFFFFF"))
