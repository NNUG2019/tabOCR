from excel2img import export_img
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from os.path import join, exists
from os import mkdir
from os import remove
import json
from skimage.io import imread, imsave
from skimage.color import rgb2gray
import numpy as np
from random import choices
from parameters_generator import (generate_table_params, COLORS, IMG_SHAPE,
                                  generate_header_params, generate_rows_height,
                                  generate_columns_width, IMG_COL_SHAPE,
                                  IMG_ROW_SHAPE)
from style_generator import (generate_border, generate_font,
                             generate_alingment, generate_pattern_fill)
from word_generator import generate_words, define_words_list
import multiprocessing as mp
import logging
import warnings
import traceback

warnings.simplefilter("ignore", UserWarning)

logger = logging.getLogger("logs")
logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(levelname)s - %(message)s')
ch.setFormatter(formatter)
logger.addHandler(ch)
fh = logging.FileHandler('logs.txt')
fh.setLevel(logging.DEBUG)
fh.setFormatter(formatter)
logger.addHandler(fh)


def convert_to_numeric(words_list):
    """ Convert numbers stored as text to number type (to avoid warning
        in Excel).
        :Arguments:
            words_list: list:: column content
        :Returns:
            list: column content (numbers stored as text converted to numbers)
    """
    if all(map(lambda s: s.isdigit(), words_list)):
        return [words_list[0]] + list(map(int, words_list[1:]))
    else:
        try:
            return [words_list[0]] + list(map(float, words_list[1:]))
        except ValueError:
            return words_list


def generate_table(words_list, table_params, font, border, alignment,
                   rows_height, column_widths):
    """ Generate .xlsx table.
        :Arguments:
            words_list: list of lists:: table content
            table_params: dict:: randomly generated table parameters
            font: Font: Font object
            border: Border:: border object
            alignment: Alignment:: alignment object
            rows_height: int:: row height
            column_widths: list:: column widths
        :Returns:
            workbook: table
    """
    workbook = Workbook()
    sheet = workbook.active
    for c in range(table_params['columns_number']):
        words = convert_to_numeric(words_list[c])
        for r in range(table_params['rows_number']):
            cell = sheet.cell(r+1, c+1)
            cell.value = words[r]
            cell.border = border
            cell.font = font
            cell.alignment = alignment
            sheet.row_dimensions[r+1].height = rows_height
            sheet.column_dimensions[get_column_letter(c+1)].width =\
                column_widths[c]
    if table_params['is_different_header']:
        header_params = generate_header_params()
        word_type = header_params['word_type']
        border = generate_border(header_params['line_style'],
                                 table_params['is_border'],
                                 table_params['is_partial_border'])
        font_header = generate_font(table_params['font_name'],
                                    header_params['font_size'],
                                    header_params['font_bold'],
                                    header_params['font_italic'])
        for c in range(table_params['columns_number']):
            cell = sheet.cell(1, c+1)
            cell.border = border
            cell.font = font_header
            cell.value = word_type(cell.value)

    return workbook


def generate_masks(columns_number, rows_number, column_widths, rows_height):
    """ Generate rgb mask for columns.
        :Arguments:
            columns_number: int:: columns number
            rows_number: int:: rows number
            column_widths: list:: column widths
            rows_height: int:: row height
        :Returns:
            workbook: table (mask)
    """
    workbook = Workbook()
    sheet = workbook.active
    for c in range(columns_number):
        for r in range(rows_number):
            cell = sheet.cell(r+1, c+1)
            sheet.row_dimensions[r+1].height = rows_height
            sheet.column_dimensions[get_column_letter(c+1)].width =\
                column_widths[c]
            cell.fill = generate_pattern_fill(COLORS[c])

    return workbook


def generate_colors(columns_number, rows_number):
    colors = [[]]
    colors_range = np.arange(0,256,1)
    for i in range(columns_number*rows_number):
        c = []
        while c in colors:
            c = choices(colors_range, k=3)
        colors.append(c)
    colors.remove([])
    colors = [y for _, y in sorted(zip(list(map(lambda c: rgb2gray(np.array([[c]], dtype='uint8')), colors)), colors))]
    colors = list(map(lambda c: ('#%02x%02x%02x' % tuple(c)).strip("#").upper(), colors))
    return list(make_chunks(colors, rows_number))


def generate_cells_mask(columns_number, rows_number, column_widths,
                        rows_height):
    colors = generate_colors(columns_number, rows_number)
    color_index = 0
    workbook = Workbook()
    sheet = workbook.active
    for c in range(columns_number):
        for r in range(rows_number):
            cell = sheet.cell(r+1, c+1)
            sheet.row_dimensions[r+1].height = rows_height
            sheet.column_dimensions[get_column_letter(c+1)].width =\
                column_widths[c]
            cell.fill = generate_pattern_fill(colors[c][r])
            color_index += 1
    return workbook


def save_data(path, table_wb, mask_wb, mask_cell_wb, words_list, table_nb, rows_number):
    """ Save .xlsx, .png and .json files for table, mask, columns,
        column content etc.
        :Arguments:
            path: str:: path to save
            table_wb: workbook:: table workbook
            mask_wb workbook:: mask workbook
            single_column_wb: list:: column workbooks
            words_list: list of lists:: table content
            table_nb: int:: table number
        :Returns:
            saved files
    """
    if not path:
        path = "dataset/"
    if not exists(path):
        mkdir(path)
    path = join(path, str(table_nb+91))
    if exists(path):
        try_again = True
        while try_again:
            try:
                remove(path)
                try_again = False
            except PermissionError:
                try_again = True
    mkdir(path)

    table_name = str(table_nb) + "_table" + ".xlsx"
    table_img_name = str(table_nb) + "_table" + ".png"
    mask_name = str(table_nb) + "_mask" + ".xlsx"
    mask_img_name = str(table_nb) + "_mask" + ".png"
    mask_cell_name = str(table_nb) + "_mask_cell" + ".xlsx"
    mask_cell_img_name = str(table_nb) + "_mask_cell" + ".png"

    table_wb.save(join(path, table_name))
    export_img(join(path, table_name), join(path, table_img_name))
    zero_padding(join(path, table_img_name), IMG_SHAPE)

    mask_wb.save(join(path, mask_name))
    export_img(join(path, mask_name), join(path, mask_img_name))
    zero_padding(join(path, mask_img_name), IMG_SHAPE, to_gray=True, rows_number=rows_number)
    remove(join(path, mask_name))

    mask_cell_wb.save(join(path, mask_cell_name))
    export_img(join(path, mask_cell_name), join(path, mask_cell_img_name))
    zero_padding(join(path, mask_cell_img_name), IMG_SHAPE, to_gray=True, rows_number=rows_number)

    img = imread(join(path, table_img_name))
    mask = imread(join(path, mask_cell_img_name))
    mask = rgb2gray(mask)
    colors = get_colors_from_mask(mask, rows_number)
    generate_column_image(img, mask, colors, words_list, table_nb, IMG_COL_SHAPE, path)
    generate_cell_image(img, mask, colors, words_list, table_nb, IMG_ROW_SHAPE, path)


def get_colors_from_mask(mask, rows_number):
    mask = rgb2gray(mask)
    colors = np.unique(mask)[1:]
    return list(make_chunks(colors, rows_number))


def generate_cell_image(img, mask, colors, words_list, table_nb, img_row_shape, path):
    for c, col_colors in enumerate(colors):
        for r, color in enumerate(col_colors):
            pos = np.where(mask == color)
            img_cell = img[pos[0].min():pos[0].max(), pos[1].min():pos[1].max(),:]

            pad_width = ((0, img_row_shape[0]-img_cell.shape[0]), (0, img_row_shape[1]-img_cell.shape[1]), (0,0))
            try:
                img_cell = np.pad(img_cell, pad_width=pad_width, mode='constant', constant_values=(0))
            except Exception as e:
                logger.error("Wrong image size: img.shape: {img.shape}, img_size: {}")
                raise

            cell_img_name = str(table_nb) + "_cell_" + str(c) + "_" + str(r) + ".png"
            cell_text_name = str(table_nb) + "_cell_" + str(c) + "_" + str(r) + ".json"
            imsave(join(path, cell_img_name), img_cell)
            with open(join(path, cell_text_name), 'w') as outfile:
                json.dump(words_list[c][r], outfile)


def generate_column_image(img, mask, colors, words_list, table_nb, img_col_shape, path):
    for c, col_colors in enumerate(colors):
        first_cell_pos = np.where(mask == col_colors[0])
        last_cell_pos = np.where(mask == col_colors[-1])
        img_col = img[first_cell_pos[0].min():last_cell_pos[0].max(),
                      first_cell_pos[1].min():last_cell_pos[1].max(),
                      :]

        pad_width = ((0, img_col_shape[0]-img_col.shape[0]), (0, img_col_shape[1]-img_col.shape[1]), (0,0))
        try:
            img_col = np.pad(img_col, pad_width=pad_width, mode='constant', constant_values=(0))
        except Exception as e:
            logger.error("Wrong image size: img.shape: {img.shape}, img_size: {}")
            raise

        col_img_name = str(table_nb) + "_column_" + str(c) + ".png"
        col_text_name = str(table_nb) + "_column_" + str(c) + ".json"
        imsave(join(path, col_img_name), img_col)
        with open(join(path, col_text_name), 'w') as outfile:
            json.dump(words_list[c], outfile)


def zero_padding(img_name, img_size, to_gray=False, rows_number=None):
    """ Image zero padding to align image shapes to defined shape, common
        to all images.
        :Arguments:
            img_name: str:: image name
            img_size: tuple:: image shape
        :Returns:
            saved, zero padded image
    """
    img = imread(img_name)
    pad = ((0, (img_size[0]-img.shape[0]+1)),
           (0, (img_size[1]-img.shape[1]+1)),
           (0, 0))
    img = img[1:, 1:, :]
    try:
        img = np.pad(img, pad_width=pad, mode='constant', constant_values=(0))
    except Exception as e:
        logger.error("Wrong image size: img.shape: {img.shape}, img_size: {}")
        raise
    if to_gray and rows_number:
        colors = get_colors_from_mask(img, rows_number)
        img = convert_rgb2gray(img, colors)
    imsave(img_name, img)


def convert_rgb2gray(img, colors):
    """ """
    # img_gray = np.zeros(img.shape[0:2])
    img_gray = rgb2gray(img)
    colors = np.unique(img_gray)
    img_new = np.zeros(img_gray.shape)
    for label, color in enumerate(colors):
        mask = img_gray == color
        mask = mask*(label+1)
        img_new += mask
    return img_new.astype('uint16')


def generate_dataset(table_nb, words_corpus, path=""):
    """ Generate single dataset, ie. .xlsx table, .png table, .png mask, .png
        columns, .json column content.
        :Arguments:
            table_nb: int:: table number
            words_corpus: dict:: words grouped by their length
            path: path to save folder
        :Returns:
            .xlsx table
            .png table
            .png mask
            .png columns
            .json column content
    """
    # for table_nb in table_nbs:  # for parallel
    # generate table parameters
    table_params = generate_table_params()
    # generate table content
    words_list = generate_words(words_corpus,
                                table_params['columns_number'],
                                table_params['rows_number'],
                                table_params['columns_type'])
    # generate column and row size
    column_widths = generate_columns_width(words_list, table_params)
    rows_height = generate_rows_height(table_params['font_size'])
    # generate table styles
    border = generate_border(table_params['line_style'],
                             table_params['is_border'],
                             table_params['is_partial_border'])
    font = generate_font(table_params['font_name'], table_params['font_size'])
    alignment = generate_alingment(table_params['horizontal_alingment'],
                                   table_params['vertical_alingment'])
    # generate table
    table_wb = generate_table(words_list, table_params, font, border,
                              alignment, rows_height, column_widths)
    mask_wb = generate_masks(table_params['columns_number'],
                             table_params['rows_number'], column_widths,
                             rows_height)
    single_columns_wb = generate_columns_images(table_wb, table_params)
    save_data(path, table_wb, mask_wb, single_columns_wb,
              words_list, table_nb)


def generate_dataset_parallel(table_nbs, words_corpus, path=""):
    """ Generate a batch of dataset, ie. .xlsx table, .png table, .png mask,
        .png columns, .json column content in parallel.
        :Arguments:
            table_nb: int:: table number
            words_corpus: dict:: words grouped by their length
            path: path to save folder
        :Returns:
            .xlsx table
            .png table
            .png mask
            .png columns
            .json column content
    """
    for table_nb in table_nbs:
        logger.info("Generate table no. {}".format(table_nb))
        should_repeat = True
        while should_repeat:
            try:
                # generate table parameters
                table_params = generate_table_params()
                # generate table content
                words_list = generate_words(words_corpus,
                                            table_params['columns_number'],
                                            table_params['rows_number'],
                                            table_params['columns_type'])
                # generate column and row size
                columns_width = generate_columns_width(words_list, table_params)
                rows_height = generate_rows_height(table_params['font_size'])
                # generate table styles
                border = generate_border(table_params['line_style'],
                                         table_params['is_border'],
                                         table_params['is_partial_border'])
                font = generate_font(table_params['font_name'],
                                     table_params['font_size'])
                alignment = generate_alingment(table_params['horizontal_alingment'],
                                               table_params['vertical_alingment'])
                # generate table
                table_wb = generate_table(words_list, table_params, font, border,
                                          alignment, rows_height, columns_width)
                mask_wb = generate_masks(table_params['columns_number'],
                                         table_params['rows_number'], columns_width,
                                         rows_height)
                mask_cell_wb =\
                    generate_cells_mask(table_params['columns_number'],
                                        table_params['rows_number'],
                                        columns_width, rows_height)
                save_data(path, table_wb, mask_wb, mask_cell_wb,
                          words_list, table_nb, table_params['rows_number'])
                should_repeat = False
                logger.info("Table was generated succesfuly")
            except Exception as e:
                tb = traceback.format_exc()
                logger.error(tb)


def make_chunks(lst, n):
    """ Split list into chunks.
        :Arguments:
            lst: list:: list to split
            n: int:: number of chunks
        :Returns:
            list of lists: list split into chunks.
    """
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def dataset_generator(number_of_tables, parallel=True, cpu_count=None,
                      save_path="", config_path=""):
    """ Generate a set of tables for learning the ANN.
        :Arguments:
            number_of_tables: int:: number of tables to generate
            parallel: bool:: is computation is made parallel
            save_path: str:: path to save
            config_path: trs:: path to config files
    """
    words_corpus = define_words_list(config_path)
    if parallel:
        # synchronous
        if not cpu_count:
            pool = mp.Pool(mp.cpu_count())
        else:
            pool = mp.Pool(cpu_count)
        chunks = make_chunks([i for i in range(number_of_tables)], 10)
        [pool.apply(generate_dataset_parallel, args=(table_nbs, words_corpus,
                                                     save_path))
         for table_nbs in chunks]
    else:
        # serialize - slow
        for table_nb in range(number_of_tables):
            generate_dataset(table_nb, words_corpus)

    # asynchronous - fast, but error can occur
    # r = [pool.apply_async(generate_dataset_parallel,
    #                      args=(table_nbs, words_corpus))
    #     for table_nbs in chunks]
    # [p.get() for p in r]
    # pool.close()
    # pool.join()


if __name__ == "__main__":
    dataset_generator(1000, cpu_count=3)
