from openpyxl import Workbook
import excel2img as e2i
import os
from colorizer import colorizer
from openpyxl.styles import PatternFill

def bare_bones_imager(lst, path):
    file_num = len(lst)

    counter = 0
    main_book = Workbook()
    main_sheet = main_book.active
    column_book = Workbook()
    column_sheet = column_book.active
    cell_book = Workbook()
    cell_sheet = cell_book.active

    colors = colorizer()
    for f in lst:
        i = 0
        for l in f:
            j = 0
            column_color = colors[i]
            column_fill = PatternFill(patternType='solid', fill_type='solid', fgColor=column_color)
            for w in l:
                cell_color = colors[17*i+j]
                cell_fill = PatternFill(patternType='solid', fill_type='solid', fgColor=cell_color)

                main_sheet.cell(row=j+1, column=i+1).value = w
                column_sheet.cell(row=j+1, column=i+1).fill = column_fill
                cell_sheet.cell(row=j+1, column=i+1).fill = cell_fill
                j += 1
            i += 1

        main_pth = os.path.join(path, 'a' + str(counter) + '.xlsx')
        main_book.save(main_pth)
        e2i.export_img(main_pth, os.path.join(path, 'a' + str(counter) + '.png'))
        column_pth = os.path.join(path, 'b' + str(counter) + '.xlsx')
        column_book.save(column_pth)
        e2i.export_img(column_pth, os.path.join(path, 'b' + str(counter) + '.png'))
        cell_pth = os.path.join(path, 'c' + str(counter) + '.xlsx')
        cell_book.save(cell_pth)
        e2i.export_img(cell_pth, os.path.join(path, 'c' + str(counter) + '.png'))
        counter += 1
        print("Copleted:", str(counter/file_num*100), "%")
