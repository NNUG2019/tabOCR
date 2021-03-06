from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from random import choice
import excel2img as e2i
import os

from colorizer import colorizer
from parameters_generator import (generate_column_params, generate_row_params,
                                  generate_border, generate_column_params,
                                  generate_column_width, generate_header_params,
                                  generate_blank_col_row_params, generate_table_params)


def full_imager(lst, path):
    """

    full_imager is major data generating function. It generates 6 files for each file string set in lst parameter:
        a<index>.xlsx - excel file containing table with strings
        a<index>.bmp - a image of table file
        b<index>.xlsx - excel file containing masked columns from table file
        b<index>.bmp - a image of column mask file
        c<index>.xlsx - excel file containing masked cells from table file
        c<index>.bmp - a image of cell mask file
    They are saved in path directory.

    :param lst: Three dimensional list of strings in format lst[file][column][row].
    :param path: Path to the target directory for saved files.
    :return:
    """
    colors = colorizer()
    counter = 0
    for f in lst:
        table_parameters = generate_table_params()

        blank_border = generate_border(line_style=table_parameters["line_style"], is_border=False)
        white_fill = PatternFill(patternType='solid', fill_type='solid', fgColor="FFFFFF")

        # Done in this way so mistakes can make more obvious crushes
        header_parameters = None
        if table_parameters["is_different_header"]:
            header_parameters = generate_header_params()

        i = 0
        row_height_sum = 0
        column_width_sum = 0
        main_book = Workbook()
        main_sheet = main_book.active
        column_book = Workbook()
        column_sheet = column_book.active
        cell_book = Workbook()
        cell_sheet = cell_book.active

        # because PyCharm doesn't like it otherwise
        j = 0
        for l in f:
            column_parameters = generate_column_params(table_parameters["columns_type"])

            j = 0
            max_word_length = 0

            # Setting column values:
            column_color = colors[i]
            column_fill = PatternFill(patternType='solid', fill_type='solid', fgColor=column_color)
            if table_parameters["is_different_header"]:
                header_column_border = generate_border(line_style=header_parameters["line_style"],
                                                       is_border=table_parameters["is_border"],
                                                       is_partial_border=table_parameters["is_partial_border"])
            text_column_border = generate_border(line_style=table_parameters["line_style"],
                                                 is_border=table_parameters["is_border"],
                                                 is_partial_border=table_parameters["is_partial_border"])
            for w in l:
                # Setting cell values:
                    # 17 so the color won't repeat!
                cell_color = colors[17*i+j]
                cell_fill = PatternFill(patternType='solid', fill_type='solid', fgColor=cell_color)

                # Row operations (one for now):
                if i == 0:
                    row_parameters = generate_row_params(font_size=table_parameters["font_size"])
                    height_of_row = row_parameters["row_height"]
                    main_sheet.row_dimensions[j+1].height = height_of_row
                    column_sheet.row_dimensions[j+1].height = height_of_row
                    cell_sheet.row_dimensions[j+1].height = height_of_row
                    row_height_sum += height_of_row

                # Generating table sheet:
                if table_parameters["is_different_header"] and j==0:
                    working_cell = main_sheet.cell(row=j+1, column=i+1)
                    working_cell.value = w
                    working_cell.border = header_column_border
                    working_cell.font = Font(name=table_parameters["font_name"],
                                             bold=header_parameters["font_bold"],
                                             italic=header_parameters["font_italic"],
                                             size=header_parameters["font_size"],
                                             color=choice(colors[1:]))
                else:
                    working_cell = main_sheet.cell(row=j+1, column=i+1)
                    working_cell.value = w
                    working_cell.border = text_column_border
                    working_cell.font = Font(name=table_parameters["font_name"],
                                             # DODO: czy dobrze rozumiem ze poza naglowkiem ma byc zawsze false ponizej?
                                             # bold=choice([True, False]),
                                             # italic=choice([True, False]),
                                             size=table_parameters["font_size"],
                                             color=choice(colors[1:]))

                # Adjusting column_width value:
                max_word_length = max(max_word_length, len(str(working_cell.value)))

                # Generating column mask sheet:
                working_cell = column_sheet.cell(row=j+1, column=i+1)
                working_cell.border = blank_border
                working_cell.fill = column_fill

                # Generating cell mask sheet:
                working_cell = cell_sheet.cell(row=j+1, column=i+1)
                working_cell.border = blank_border
                working_cell.fill = cell_fill

                j += 1

            # Setting correct column widths in sheets:
            width_of_column = generate_column_width(font_size=table_parameters["font_size"],
                                                    max_word_length=max_word_length)["column_width"]
            main_sheet.column_dimensions[get_column_letter(i+1)].width = width_of_column
            column_sheet.column_dimensions[get_column_letter(i+1)].width = width_of_column
            cell_sheet.column_dimensions[get_column_letter(i+1)].width = width_of_column
            column_width_sum += width_of_column
            i += 1

        # Creating blank rows and columns:
        blank_row_height, blank_column_width = generate_blank_col_row_params(rows_height=row_height_sum,
                                                                             columns_width=column_width_sum)
        #main_sheet.cell(row=j + 1, column=i + 1).value = "0"
        #column_sheet.cell(row=j + 1, column=i + 1).value = "0"
        #cell_sheet.cell(row=j + 1, column=i + 1).value = "0"

        for n in range(j):
            main_sheet.cell(row=n + 1, column=i + 1).border = blank_border
            main_sheet.cell(row=n + 1, column=i + 1).fill = white_fill
            cell_sheet.cell(row=n + 1, column=i + 1).border = blank_border
            cell_sheet.cell(row=n + 1, column=i + 1).fill = white_fill
            column_sheet.cell(row=n + 1, column=i + 1).border = blank_border
            column_sheet.cell(row=n + 1, column=i + 1).fill = white_fill

        for k in range(i):
            main_sheet.cell(row=j + 1, column=k + 1).border = blank_border
            main_sheet.cell(row=j + 1, column=k + 1).fill = white_fill
            cell_sheet.cell(row=j + 1, column=k + 1).border = blank_border
            cell_sheet.cell(row=j + 1, column=k + 1).fill = white_fill
            column_sheet.cell(row=j + 1, column=k + 1).border = blank_border
            column_sheet.cell(row=j + 1, column=k + 1).fill = white_fill


        main_sheet.column_dimensions[get_column_letter(i+1)].width = blank_column_width
        main_sheet.row_dimensions[j+1].height = blank_row_height
        column_sheet.column_dimensions[get_column_letter(i+1)].width = blank_column_width
        column_sheet.row_dimensions[j+1].height = blank_row_height
        cell_sheet.column_dimensions[get_column_letter(i+1)].width = blank_column_width
        cell_sheet.row_dimensions[j+1].height = blank_row_height


        # Saving table sheet:
        main_pth = os.path.join(path, 'a' + str(counter) + '.xlsx')
        main_book.save(main_pth)
        e2i.export_img(main_pth, os.path.join(path, 'a' + str(counter) + '.bmp'))

        # Saving column mask sheet:
        column_pth = os.path.join(path, 'b' + str(counter) + '.xlsx')
        column_book.save(column_pth)
        e2i.export_img(column_pth, os.path.join(path, 'b' + str(counter) + '.bmp'))

        # Saving cell mask sheet:
        cell_pth = os.path.join(path, 'c' + str(counter) + '.xlsx')
        cell_book.save(cell_pth)
        e2i.export_img(cell_pth, os.path.join(path, 'c' + str(counter) + '.bmp'))

        counter += 1
