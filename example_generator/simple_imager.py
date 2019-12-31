from openpyxl import Workbook
import excel2img as e2i
import os
from colorizer import colorizer
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from random import choice

def simple_imager(lst, path):
    file_num = len(lst)

    counter = 0
    main_book = Workbook()
    main_sheet = main_book.active
    column_book = Workbook()
    column_sheet = column_book.active
    cell_book = Workbook()
    cell_sheet = cell_book.active


    def take(mods):
        mod = choice(mods)
        mods.remove(mod)
        return mod

    colors = colorizer()
    for f in lst:

        # Bardzo brzydkie rozwiazanie na szybko.

        col_colors = colorizer()
        cell_colors = col_colors.copy()

        modi1 = [-2, -3, -1, -1, -1, 1, 2, 3, 1, 1]
        modi2 = modi1.copy()
        modi3 = modi1.copy()

        modi4 = [-2, -3, -1, 0, -1, 1, 3, 2, 0, 1]
        modi5 = modi4.copy()
        modi6 = modi4.copy()

        i = 0
        for l in f:
            j = 0

            column_color = take(col_colors)
            # colors[i] zrobi gradient
            column_fill = PatternFill(patternType='solid', fill_type='solid', fgColor=column_color)
            for w in l:
                cell_color = take(cell_colors)
                # colors[17*i+j] zrobi gradient
                cell_fill = PatternFill(patternType='solid', fill_type='solid', fgColor=cell_color)

                main_sheet.cell(row=j+1, column=i+1).value = w
                column_sheet.cell(row=j+1, column=i+1).fill = column_fill
                cell_sheet.cell(row=j+1, column=i+1).fill = cell_fill
                j += 1
            i += 1

            letter = get_column_letter(i)
            main_sheet.column_dimensions[letter].width = 10 + take(modi1)
            column_sheet.column_dimensions[letter].width = 10 + take(modi2)
            cell_sheet.column_dimensions[letter].width = 10 + take(modi3)

            main_sheet.row_dimensions[i].height = 25 + 2*take(modi4)
            column_sheet.row_dimensions[i].height = 25 + 2*take(modi5)
            cell_sheet.row_dimensions[i].height = 25 + 2*take(modi6)

        main_pth = os.path.join(path, 'a' + str(counter) + '.xlsx')
        main_book.save(main_pth)
        e2i.export_img(main_pth, os.path.join(path, 'a' + str(counter) + '.png'))
        column_pth = os.path.join(path, 'b' + str(counter) + '.xlsx')
        column_book.save(column_pth)
        e2i.export_img(column_pth, os.path.join(path, 'b' + str(counter) + '.png'))
        cell_pth = os.path.join(path, 'c' + str(counter) + '.xlsx')
        cell_book.save(cell_pth)
        e2i.export_img(cell_pth, os.path.join(path, 'c' + str(counter) + '.png'))
        counter += 1
        print("Copleted:", str(counter/file_num*100), "%")
